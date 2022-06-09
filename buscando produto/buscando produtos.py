import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://lista.mercadolivre.com.br/'

tabela_inicial = openpyxl.load_workbook('produtos mercado livre.xlsx')
primeira_aba = tabela_inicial.active

for celula in primeira_aba['B']:
    produto_nome = celula.value

resposta = requests.get(url + produto_nome)

site = BeautifulSoup(resposta.text, 'html.parser')

produtos = site.findAll('div', attrs={'class': 'andes-card andes-card--flat andes-card--default ui-search-result ui-search-result--core andes-card--padding-default'})
print(produtos)

tabela_produto = openpyxl.Workbook()
tabela_produto.create_sheet(produto_nome)
info_produto = tabela_produto[produto_nome]
info_produto.append(['TITULO', 'LINK', 'PREÇO', 'DESCONTO'])

for produto in produtos:

    link = produto.find('a', attrs={'class': 'ui-search-link'})
    titulo = produto.find('h2', attrs={'class': 'ui-search-item__title'})

    real = produto.find('span', attrs={'class': 'price-tag-fraction'})
    centavos = produto.find('span', attrs={'class': 'price-tag-cents'})

    desconto = produto.find('span', attrs={'class': 'ui-search-price__discount'})

    print('Título do produto:', titulo.text)
    print('Link do produto:', link['href'])

    if centavos:
        print('Esse é o valor:', real.text + ',' + centavos.text)
    else:
        print('Esse é o valor:', real.text + ',00' )

    if desconto:
        print('Esta com desconto de:', desconto.text)
    
    if desconto and centavos:
        info_produto.append([titulo.text, link['href'], real.text + ',' + centavos.text, desconto.text])
    if desconto and not centavos:
        info_produto.append([titulo.text, link['href'], real.text + ',00', desconto.text])
    if centavos and not desconto:
        info_produto.append([titulo.text, link['href'], real.text + ',' + centavos.text])
    if not centavos and not desconto:
        info_produto.append([titulo.text, link['href'], real.text + ',00'])
    
    print('\n\n')

info_produto = tabela_produto['Sheet']
info_produto.append(['produto -->'])

tabela_produto.save('produtos mercado livre.xlsx')